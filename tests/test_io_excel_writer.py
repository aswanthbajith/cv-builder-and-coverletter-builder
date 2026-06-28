"""Tests for the v2 Excel writer (C5 wiring)."""

from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path

import pytest

from job_automation.io.excel_writer import V2_COLUMNS, write_generation_results
from job_automation.models.job import Job
from job_automation.models.results import GenerationResult, MatchResult


def _make_result(
    company: str,
    title: str,
    *,
    score: float = 75.0,
    tex_path: Path | None = None,
    pdf_path: Path | None = None,
    error: str | None = None,
) -> GenerationResult:
    return GenerationResult(
        job=Job(
            company=company,
            job_title=title,
            location="Remote",
            job_description="desc",
            required_skills="Python",
            preferred_qualifications="",
        ),
        match=MatchResult(
            overall_score=score,
            education_score=50.0,
            skills_score=score,
            programming_score=score,
            research_score=50.0,
            experience_score=score,
            reasoning="matched=3, partial=1, missing=2",
            missing_skills=["Rust"],
            strengths=["Python"],
            status="review",
        ),
        resume_tex_path=tex_path,
        resume_pdf_path=pdf_path,
        cover_letter_path=None,
        generated_at=datetime(2026, 6, 28, tzinfo=timezone.utc),
        error=error,
    )


def test_write_generation_results_creates_new_file(tmp_path: Path) -> None:
    """When no file exists, write_generation_results creates one with v2 columns."""
    output = tmp_path / "Job_Matching_Updated.xlsx"
    tex = tmp_path / "acme.tex"
    pdf = tmp_path / "acme.pdf"
    results = [_make_result("Acme", "HPC Engineer", tex_path=tex, pdf_path=pdf)]
    written = write_generation_results(results, output)
    assert written == 1
    assert output.exists()

    import openpyxl

    wb = openpyxl.load_workbook(output)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header, *body = rows
    assert header == tuple(V2_COLUMNS)
    assert len(body) == 1
    row = dict(zip(header, body[0]))
    assert str(row["Company"]) == "Acme"
    assert str(row["Job_Title"]) == "HPC Engineer"
    assert str(row["_engine"]) == "v2"
    assert float(row["match_percent"]) == 75.0
    assert "HYPERLINK" in str(row["resume_pdf"])
    assert "acme.pdf" in str(row["resume_pdf"])


def test_write_generation_results_preserves_legacy_rows(tmp_path: Path) -> None:
    """Legacy rows (no _engine marker) survive a v2 run."""
    output = tmp_path / "Job_Matching_Updated.xlsx"

    # First, plant a legacy file with one legacy row.
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Job Matches"
    legacy_columns = list(V2_COLUMNS) + ["legacy_only_column"]
    ws.append(legacy_columns)
    ws.append(
        [
            "Legacy Co",
            "Legacy Title",
            60.0,
            "legacy_reasoning",
            "",
            "",
            "review",
            "",
            "",
            "",
            "2026-01-01T00:00:00+00:00",
            "",
            "legacy_extra_value",
        ]
    )
    wb.save(output)

    # Now run the v2 writer with a different job.
    results = [_make_result("V2 Co", "V2 Title", score=88.0)]
    written = write_generation_results(results, output)
    assert written == 1

    wb = openpyxl.load_workbook(output)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header, *body = rows
    assert "legacy_only_column" in header
    # First body row is the legacy row, second is the v2 row.
    legacy_row = dict(zip(header, body[0]))
    v2_row = dict(zip(header, body[1]))
    assert legacy_row["Company"] == "Legacy Co"
    assert legacy_row["legacy_only_column"] == "legacy_extra_value"
    assert v2_row["Company"] == "V2 Co"
    assert v2_row["_engine"] == "v2"


def test_write_generation_results_overwrites_in_place(tmp_path: Path) -> None:
    """Existing v2 rows for matching (Company, Job_Title) are overwritten."""
    output = tmp_path / "Job_Matching_Updated.xlsx"

    # Plant an existing v2 row for Acme/HPC Engineer.
    results_initial = [_make_result("Acme", "HPC Engineer", score=50.0)]
    write_generation_results(results_initial, output)

    # Re-run with a different score for the same (Company, Title).
    results_updated = [_make_result("Acme", "HPC Engineer", score=95.0)]
    write_generation_results(results_updated, output)

    import openpyxl

    wb = openpyxl.load_workbook(output)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header, *body = rows
    assert len(body) == 1, "v2 row must be overwritten in place, not duplicated"
    row = dict(zip(header, body[0]))
    assert row["Company"] == "Acme"
    assert float(row["match_percent"]) == 95.0


def test_write_generation_results_handles_none_paths(tmp_path: Path) -> None:
    """No tex/pdf path → empty cells (no error)."""
    output = tmp_path / "Job_Matching_Updated.xlsx"
    results = [_make_result("Co", "Title", tex_path=None, pdf_path=None)]
    written = write_generation_results(results, output)
    assert written == 1

    import openpyxl

    wb = openpyxl.load_workbook(output)
    ws = wb.active
    header, *body = list(ws.iter_rows(values_only=True))
    row = dict(zip(header, body[0]))
    assert row["resume_tex"] in (None, "")
    assert row["resume_pdf"] in (None, "")


def test_write_generation_results_appends_new_rows(tmp_path: Path) -> None:
    """Multiple distinct results are appended; legacy rows are preserved."""
    output = tmp_path / "Job_Matching_Updated.xlsx"

    # Plant a legacy file.
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Job Matches"
    ws.append(list(V2_COLUMNS))
    ws.append(["Legacy", "Title", 60.0, "r", "", "", "review", "", "", "", "2026", ""])
    wb.save(output)

    # Run v2 with two new jobs.
    results = [
        _make_result("CoA", "TA"),
        _make_result("CoB", "TB"),
    ]
    written = write_generation_results(results, output)
    assert written == 2

    wb = openpyxl.load_workbook(output)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header, *body = rows
    assert len(body) == 3
    companies = [dict(zip(header, r))["Company"] for r in body]
    assert companies == ["Legacy", "CoA", "CoB"]


def test_write_generation_results_empty_input(tmp_path: Path) -> None:
    """Empty input creates the file with just the header row."""
    output = tmp_path / "Job_Matching_Updated.xlsx"
    written = write_generation_results([], output)
    assert written == 0
    assert output.exists()

    import openpyxl

    wb = openpyxl.load_workbook(output)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    assert len(rows) == 1
    assert list(rows[0]) == list(V2_COLUMNS)