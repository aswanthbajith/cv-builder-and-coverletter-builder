"""Excel writer for v2 pipeline results.

Reads ``output/Job_Matching_Updated.xlsx`` (if it exists), preserves
legacy rows that aren't v2 results, and overwrites v2 rows in place.
New v2 rows that have no legacy counterpart are appended at the bottom.

The output schema is a deliberate superset of the legacy schema
(``Company``, ``Job_Title``, ``match_percent``, ``match_reasoning``,
``missing_skills``, ``strengths``, ``status``, ``resume_tex``,
``resume_pdf``, ``cover_letter``, ``generated_date``,
``_engine``). The ``_engine`` column is the marker used to identify
v2 rows for in-place overwrite.

HYPERLINK() formulas are re-emitted on every write so that openpyxl
preserves them as live links.
"""

from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path

from job_automation.models.results import GenerationResult

logger = logging.getLogger(__name__)


V2_COLUMNS: list[str] = [
    "Company",
    "Job_Title",
    "match_percent",
    "match_reasoning",
    "missing_skills",
    "strengths",
    "status",
    "resume_tex",
    "resume_pdf",
    "cover_letter",
    "generated_date",
    "_engine",
]


def _hyperlink_formula(relative_path: Path | str | None, label: str) -> str:
    """Return an Excel HYPERLINK() formula for a relative path.

    Excel renders HYPERLINK("path", "label") as a clickable cell. We
    use forward slashes because openpyxl preserves formulas as text
    and Excel on Windows handles either slash style.
    """
    if relative_path is None:
        return ""
    posix_path = Path(relative_path).as_posix()
    return f'=HYPERLINK("{posix_path}","{label}")'


def _row_from_result(result: GenerationResult) -> dict[str, object]:
    """Project a :class:`GenerationResult` into a v2 Excel row."""
    out: dict[str, object] = {
        "Company": result.job.company,
        "Job_Title": result.job.job_title,
        "match_percent": result.match.overall_score,
        "match_reasoning": result.match.reasoning,
        "missing_skills": ", ".join(result.match.missing_skills),
        "strengths": ", ".join(result.match.strengths),
        "status": result.match.status,
        "resume_tex": _hyperlink_formula(
            _relative_to_root(result.resume_tex_path), "Open Tex"
        ),
        "resume_pdf": _hyperlink_formula(
            _relative_to_root(result.resume_pdf_path), "Open PDF"
        ),
        "cover_letter": "",  # cover letter is H1, not C5
        "generated_date": result.generated_at.isoformat()
        if isinstance(result.generated_at, datetime)
        else str(result.generated_at),
        "_engine": "v2",
    }
    if result.error:
        # Surface pipeline-level errors so users can spot them.
        out["match_reasoning"] = (
            f"{result.match.reasoning} [error: {result.error}]"
        )
    return out


def _relative_to_root(path: Path | None) -> Path | None:
    """Best-effort path relative to cwd; fall back to absolute."""
    if path is None:
        return None
    try:
        return path.resolve().relative_to(Path.cwd().resolve())
    except ValueError:
        return path


def _load_existing_rows(output_path: Path) -> tuple[list[dict[str, object]], list[str]]:
    """Load existing rows from the Excel sheet (returns ``([], [])`` on miss)."""
    if not output_path.exists():
        return [], list(V2_COLUMNS)
    try:
        import openpyxl

    except ImportError:
        logger.warning("openpyxl_missing_skipping_excel_load")
        return [], list(V2_COLUMNS)

    wb = openpyxl.load_workbook(output_path)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        return [], list(V2_COLUMNS)

    columns: list[str] = [str(h) if h else f"col_{i}" for i, h in enumerate(header)]
    rows: list[dict[str, object]] = []
    for raw in rows_iter:
        if raw is None or all(v is None for v in raw):
            continue
        rows.append({columns[i]: raw[i] for i in range(min(len(columns), len(raw)))})
    return rows, columns


def write_generation_results(
    results: list[GenerationResult],
    output_path: Path,
    *,
    preserve_legacy_rows: bool = True,
) -> int:
    """Write v2 results to ``output_path``, preserving legacy rows.

    Returns the number of v2 rows written (new + overwritten).
    """
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover - openpyxl is a hard dep
        raise RuntimeError("openpyxl is required for excel_writer") from exc

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    existing_rows, existing_columns = _load_existing_rows(output_path)

    # Build v2 row index by (Company, Job_Title). A given (company, title)
    # pair always has the same v2 result on a given run, so we overwrite.
    v2_keys: set[tuple[str, str]] = set()
    for result in results:
        v2_keys.add((result.job.company, result.job.job_title))

    new_rows: list[dict[str, object]] = []
    if preserve_legacy_rows:
        for row in existing_rows:
            # Preserve the row if it isn't a v2 row matching one of our keys.
            engine_marker = str(row.get("_engine", "") or "")
            company = str(row.get("Company", "") or "")
            title = str(row.get("Job_Title", "") or "")
            if engine_marker == "v2" and (company, title) in v2_keys:
                continue  # will be replaced below
            new_rows.append(row)

    written = 0
    for result in results:
        new_rows.append(_row_from_result(result))
        written += 1

    # Merge columns: union of V2_COLUMNS and any preserved legacy columns.
    final_columns: list[str] = list(V2_COLUMNS)
    if preserve_legacy_rows:
        for col in existing_columns:
            if col and col not in final_columns:
                final_columns.append(col)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Job Matches"
    ws.append(final_columns)
    for row in new_rows:
        ws.append([row.get(col, "") for col in final_columns])

    # Auto-size columns (capped) for readability.
    for idx, col in enumerate(final_columns, start=1):
        max_len = len(str(col))
        for row in new_rows:
            cell_value = row.get(col, "")
            cell_str = "" if cell_value is None else str(cell_value)
            if len(cell_str) > max_len:
                max_len = len(cell_str)
        ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = min(
            max_len + 2, 60
        )

    wb.save(output_path)
    logger.info(
        "excel_written",
        extra={"path": str(output_path), "rows": written, "total": len(new_rows)},
    )
    return written


__all__ = ["V2_COLUMNS", "write_generation_results"]
